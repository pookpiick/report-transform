#!/usr/bin/env python3
"""
Read CSV files from input/ (columns: Page, Text) and write one Excel file per CSV
to output/, using the template structure with Page -> Page. and Text -> OE/Owner Comment.
"""

import csv
import io
import sys
from pathlib import Path

import openpyxl

# Paths relative to project root
PROJECT_ROOT = Path(__file__).resolve().parent
INPUT_DIR = PROJECT_ROOT / "input"
OUTPUT_DIR = PROJECT_ROOT / "output"
TEMPLATE_PATH = OUTPUT_DIR / "comment_response_template.xlsx"

# Template: row 1 = title, row 2 = headers
# Column A = "No.", Column B = "Rev.", Column C = "Page.", Column D = "OE/Owner Comment"
HEADER_ROW = 2
DATA_START_ROW = 3
COL_NO = 1
COL_REV = 2
COL_PAGE = 3
COL_OE_OWNER_COMMENT = 4


def get_input_files() -> list[Path]:
    """Return all CSV files in input directory."""
    if not INPUT_DIR.is_dir():
        return []
    return sorted(INPUT_DIR.glob("*.csv"))


def read_previous_revision_rows(workbook_file) -> list[list]:
    """
    Read all data rows from a previous revision Excel file, preserving every column.
    workbook_file: path or file-like (bytes). Returns list of rows; each row is a list
    of cell values for columns 1 to max_column (so previous file's other columns are kept).
    """
    if isinstance(workbook_file, (str, Path)):
        wb = openpyxl.load_workbook(workbook_file)
    else:
        workbook_file.seek(0)
        wb = openpyxl.load_workbook(io.BytesIO(workbook_file.read()))
    ws = wb.active
    max_col = ws.max_column
    rows = []
    r = DATA_START_ROW
    while True:
        no_val = ws.cell(row=r, column=COL_NO).value
        page_val = ws.cell(row=r, column=COL_PAGE).value
        comment_val = ws.cell(row=r, column=COL_OE_OWNER_COMMENT).value
        if no_val is None and page_val is None and comment_val is None:
            break
        row_cells = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        rows.append(row_cells)
        r += 1
    return rows


def transform_csv_to_workbook(csv_file, template_path: Path, revision: str | None = None, previous_revision_file=None):
    """
    Read CSV (Page, Text) from file-like csv_file, fill template workbook, return workbook.
    csv_file: path or file-like object (text stream) with UTF-8 CSV data.
    revision: optional value (e.g. A–G, 0–4) to write into the "Rev." column for new records.
    previous_revision_file: optional path or file-like (bytes) of a previous Excel output; its rows are prepended.
    """
    if isinstance(csv_file, (str, Path)):
        f = open(csv_file, newline="", encoding="utf-8-sig")
        try:
            return _fill_workbook_from_csv(f, template_path, revision, previous_revision_file)
        finally:
            f.close()
    if hasattr(csv_file, "seek"):
        csv_file.seek(0)
    return _fill_workbook_from_csv(csv_file, template_path, revision, previous_revision_file)


def _fill_workbook_from_csv(csv_stream, template_path: Path, revision: str | None = None, previous_revision_file=None):
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    previous_rows = []
    if previous_revision_file is not None:
        previous_rows = read_previous_revision_rows(previous_revision_file)

    reader = csv.DictReader(csv_stream)
    if "Page" not in reader.fieldnames or "Text" not in reader.fieldnames:
        raise ValueError(
            f"CSV must have columns 'Page' and 'Text'; got {reader.fieldnames}"
        )
    SKIP_TEXT_VALUES = frozenset(s.strip().lower() for s in ("Closed", "Close", "Open", "Opened"))
    new_rows = [
        row for row in reader
        if (row.get("Text") or "").strip() and (row.get("Text") or "").strip().lower() not in SKIP_TEXT_VALUES
    ]

    row_index = 0
    for prev_cells in previous_rows:
        r = DATA_START_ROW + row_index
        row_index += 1
        for col_idx, value in enumerate(prev_cells, start=1):
            if col_idx == COL_NO:
                ws.cell(row=r, column=col_idx, value=row_index)
            else:
                ws.cell(row=r, column=col_idx, value=value)

    for row in new_rows:
        r = DATA_START_ROW + row_index
        row_index += 1
        page_val = row.get("Page") or row.get("\ufeffPage")
        text_val = row.get("Text")
        ws.cell(row=r, column=COL_NO, value=row_index)
        ws.cell(row=r, column=COL_REV, value=revision)
        ws.cell(row=r, column=COL_PAGE, value=page_val)
        ws.cell(row=r, column=COL_OE_OWNER_COMMENT, value=text_val)

    return wb


def transform_csv_to_xlsx(csv_path: Path, output_path: Path, template_path: Path, revision: str | None = None) -> None:
    """
    Read CSV (Page, Text), fill template workbook, save to output_path.
    """
    wb = transform_csv_to_workbook(csv_path, template_path, revision)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    if not TEMPLATE_PATH.exists():
        print(f"Template not found: {TEMPLATE_PATH}", file=sys.stderr)
        return 1

    input_files = get_input_files()
    if not input_files:
        print(f"No CSV files found in {INPUT_DIR}", file=sys.stderr)
        return 1

    for csv_path in input_files:
        # Output name: same base name as input, with .xlsx
        output_name = csv_path.stem + ".xlsx"
        output_path = OUTPUT_DIR / output_name
        try:
            transform_csv_to_xlsx(csv_path, output_path, TEMPLATE_PATH)
            print(f"Created {output_path}")
        except Exception as e:
            print(f"Error processing {csv_path}: {e}", file=sys.stderr)
            return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
